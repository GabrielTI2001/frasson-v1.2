from django.shortcuts import render, HttpResponse
from rest_framework import permissions, viewsets, status
from rest_framework.response import Response
from django.http import JsonResponse
from django.db.models import Q
from rest_framework.permissions import IsAuthenticated
import requests, json, time, os
from .serializers import *
from .models import Operacoes_Contratadas, Operacoes_Contratadas_Cedulas, Operacoes_Contratadas_Glebas
from io import BytesIO
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment
from django.views.decorators.csrf import csrf_exempt
from pykml import parser
from lxml import etree
from services.views import parse_element_kml
from pykml.factory import KML_ElementMaker as KML


class ItensFinanciadosView(viewsets.ModelViewSet):
    queryset = Itens_Financiados.objects.all()
    serializer_class = listItemsFinanciados
    lookup_field = 'uuid'

class OperacoesContratadasView(viewsets.ModelViewSet):
    queryset = Operacoes_Contratadas.objects.all()
    serializer_class = detailOperacoes
    lookup_field = 'uuid'
    # permission_classes = [permissions.AllowAny]
    def get_queryset(self):
        queryset = super().get_queryset()
        search = self.request.query_params.get('search', None) 
        search_year = self.request.query_params.get('year', None)   
        search_month = self.request.query_params.get('month', None)  
        search_bank = self.request.query_params.get('bank', None)  
        query = Q()
        if search_year:
            query &= Q(data_emissao_cedula__year=search_year)
            if search_month:
                query &= Q(data_emissao_cedula__month=search_month)
        if search_bank:
            query &= Q(instituicao__instituicao_id=search_bank)      
        if search:
            query &= Q(beneficiario__razao_social__icontains=search)
        queryset = queryset.filter(query).order_by('beneficiario__razao_social', 'data_emissao_cedula')
        return queryset
    def get_serializer_class(self):
        if self.action == 'list':
            return listOperacoes
        else:
            return self.serializer_class
    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        docs = request.FILES.getlist('file')
        kml = request.FILES.get('kml')
        if serializer.is_valid():
            docs_validos = 0
            for i in docs:
                ext = os.path.splitext(i.name)[1]
                valid_extensions = ['.pdf']
                if not ext.lower() in valid_extensions:
                    return Response({'file': 'Os arquivos precisa, ser em formato PDF!'}, status=400)
                filesize = i.size
                if filesize > 8 * 1024 * 1024:  # 2 MB
                    return Response({'file': 'O tamanho máximo do arquivo é 8 MB!'}, status=400)
                docs_validos += 1
            if docs_validos == len(docs):
                self.perform_create(serializer)
                for i in docs:
                    Operacoes_Contratadas_Cedulas.objects.create(operacao=serializer.instance, upload_by=request.user, file=i)
            if kml:
                ext = os.path.splitext(kml.name)[1]
                valid_extensions = ['.kml']
                if not ext.lower() in valid_extensions:
                    return Response({'kml': 'O arquivo precisa ser em formato KML!'}, status=400)
                operacao = serializer.save()
                kml.seek(0) 
                root = parser.parse(kml).getroot()
                coordinates_gd = parse_element_kml(root)
                for latlong in coordinates_gd:
                    Operacoes_Contratadas_Glebas.objects.create(operacao_id=operacao.id, latitude_gd=latlong['lat'], longitude_gd=latlong['lng'])
                headers = self.get_success_headers(serializer.data)
                return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)
            else:
                return Response({'kml': 'Submeta um Arquivo!'}, status=400)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        if serializer.is_valid():
            operacao = serializer.save()
            kml = request.FILES.get('kml')
            if kml:
                ext = os.path.splitext(kml.name)[1]
                valid_extensions = ['.kml']
                if not ext.lower() in valid_extensions:
                    return Response({'kml': 'O arquivo precisa ser em formato KML!'}, status=400)
                kml.seek(0) 
                root = parser.parse(kml).getroot()
                coordinates_gd = parse_element_kml(root)
                Operacoes_Contratadas_Glebas.objects.filter(regime_id=operacao.id).delete()
                for latlong in coordinates_gd:
                    Operacoes_Contratadas_Glebas.objects.create(operacao_id=operacao.id, latitude_gd=latlong['lat'], longitude_gd=latlong['lng'])
            headers = self.get_success_headers(serializer.data)
            return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class OperacoesCedulasView(viewsets.ModelViewSet):
    queryset = Operacoes_Contratadas_Cedulas.objects.all()
    serializer_class = serOperacoesCedulas
    def create(self, request, *args, **kwargs):
        response_data = []
        serializer = self.get_serializer(data=request.data)
        images = request.FILES.getlist('file')
        if serializer.is_valid():
            imgs_validas = 0
            for i in images:
                ext = os.path.splitext(i.name)[1]
                valid_extensions = ['.pdf']
                if not ext.lower() in valid_extensions:
                    return Response({'error': 'O arquivo precisa ser em formato PDF!'}, status=400)
                filesize = i.size
                if filesize > 8 * 1024 * 1024:  # 8 MB
                    return Response({'error': 'O tamanho máximo da imagem é 8 MB!'}, status=400)
                imgs_validas += 1
            if imgs_validas == len(images):
                for i in images:
                    reg = Operacoes_Contratadas_Cedulas.objects.create(operacao_id=request.data.get('operacao'), upload_by=request.user, file=i)
                    response_data.append({'id':reg.id, 'url':'/media/'+reg.file.name})
                headers = self.get_success_headers(serializer.data)
                return Response(response_data, status=status.HTTP_201_CREATED, headers=headers)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        if serializer.is_valid():
            ext = os.path.splitext(request.FILES.get('file').name)[1]
            valid_extensions = ['.pdf']
            if not ext.lower() in valid_extensions:
                return Response({'error': 'O arquivo precisa ser em formato PDF!'}, status=400)
            filesize = request.FILES.get('file').size
            if filesize > 8 * 1024 * 1024:  # 8 MB
                return Response({'error': 'O tamanho máximo do arquivo é 8 MB!'}, status=400)
            else:
                serializer.save()
                headers = self.get_success_headers(serializer.data)
                return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


def creditData(request):
    anos_operacoes = [y['data_emissao_cedula__year'] for y in Operacoes_Contratadas.objects.values('data_emissao_cedula__year').distinct()]
    instituicoes = Operacoes_Contratadas.objects.values('instituicao__instituicao__id','instituicao__instituicao__razao_social').distinct()
    lista_instituicoes = [{
        'id': instituicao['instituicao__instituicao__id'],
        'instituicao': instituicao['instituicao__instituicao__razao_social']
    }for instituicao in instituicoes]

    data = {
        'anos': anos_operacoes,
        'instituicoes': lista_instituicoes,
    }
    return JsonResponse(data)


@csrf_exempt
def convert_html_table_to_excel(request):
    if request.method == "POST":
        time_now = int(time.time())
        file_name = f"report_{time_now}"
        request_data = json.loads(request.body)
        html_content = request_data.get('html_content', '')
        soup = BeautifulSoup(html_content, 'html.parser')
        table = soup.find('table')

        # Cria um Excel Workbook
        wb = Workbook()
        ws = wb.active  # Get active worksheet

        col_widths = {}  # A dictionary to hold the maximum length of each column
        center_aligned_style = Alignment(horizontal='center', vertical='center')
        for i, row in enumerate(table.find_all('tr')):
            for j, cell in enumerate(row.find_all(['td', 'th'])):
                cell_value = cell.get_text().strip()
                # Prrenche colunas do excel com valores da table
                current_cell = ws.cell(row=i+1, column=j+1, value=cell_value)
                current_cell.alignment = center_aligned_style 
                col_widths[j] = max(col_widths.get(j, 0), len(cell_value) + 2)

        # Apply column widths
        for idx, width in col_widths.items():
            col_letter = ws.cell(row=1, column=idx+1).column_letter
            ws.column_dimensions[col_letter].width = width

        # Create a BytesIO object and save the workbook to it
        excel_file = BytesIO()
        wb.save(excel_file)
        # Set the pointer of BytesIO object to the beginning
        excel_file.seek(0)

        response = HttpResponse(
            excel_file.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename={file_name}.xlsx'
        excel_file.close()

        return response
    else:
        return HttpResponse(404)

def download_kml_operacao(request, uuid):
    time_now = int(time.time()) #for the file name
    file_name = f"imovel_{time_now}.kml"
    data = Operacoes_Contratadas_Glebas.objects.values().filter(operacao__uuid=uuid)

    polygon_coordinates = []
    for coord in data:
        str_coordenada = f"{coord['longitude_gd']},{coord['latitude_gd']}"
        polygon_coordinates.append(str_coordenada)

    first_coordinate = f"{data[0]['longitude_gd']},{data[0]['latitude_gd']}" #first coordinate
    polygon_coordinates.append(first_coordinate)  #add again the first coordinate to close the polygon

    # Create a Polygon Style
    polygon_style = KML.Style(
        KML.id("polygon_style"),
        KML.PolyStyle(
            KML.color('00ffffff'),  # Fully transparent
            KML.fill(0),  # 0 means no fill
        ),
        KML.LineStyle(
            KML.color('ff00ff00'),  # Green outline
            KML.width(2),  # You can also set the line width
        )
    )

    # Create a KML Document with a Polygon and Style
    kml_doc = KML.kml(
        KML.Document(
            polygon_style,  # Add the Style to the Document
            KML.Folder(
                KML.name("Polygon"),
                KML.Placemark(
                    KML.name("Polygon"),
                    KML.styleUrl("#polygon_style"),  # Refer to the polygon_style by id
                    KML.Polygon(
                        KML.outerBoundaryIs(
                            KML.LinearRing(
                                KML.coordinates(" ".join(polygon_coordinates))
                            )
                        )
                    )
                )
            )
        )
    )
    # Convert KML document to string
    kml_str = etree.tostring(kml_doc, pretty_print=True)
    # Create a response with the KML type
    response = HttpResponse(kml_str, content_type='application/vnd.google-earth.kml+xml')
    # Add a file attachment header
    response['Content-Disposition'] = f'attachment; filename="{file_name}"'
    return response
