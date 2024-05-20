from django.shortcuts import render
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from .serializers import serCommodity_Prices, serLocations, serProdutos
from rest_framework import viewsets
from .models import Commodity_Prices, Localizacao_Cotacao, Commodity_Cotacao
from django.db.models import Q
from django.db.models import Avg, StdDev, Min, Max
from rest_framework.response import Response
from backend.frassonUtilities import Frasson
from pykml import parser
from lxml import etree
from io import BytesIO
from backend.settings import TOKEN_GOOGLE_MAPS_API
import time, math
from bs4 import BeautifulSoup
from .tasks import update_commodity_prices
from django.http import FileResponse, HttpResponse
import io, json
from openpyxl import Workbook
from openpyxl.styles import Alignment
from pykml.factory import KML_ElementMaker as KML

# Create your views here.
class CommodityPriceView(viewsets.ModelViewSet):
    queryset = Commodity_Prices.objects.all()
    serializer_class = serCommodity_Prices
    def get_queryset(self):
        queryset = super().get_queryset()
        produto = self.request.query_params.get('produto', None)   
        local = self.request.query_params.get('local', None)  
        inicio = self.request.query_params.get('inicio', None) 
        final = self.request.query_params.get('final', None) 

        if produto and local and inicio and final:
            queryset = queryset.filter(Q(Q(commodity_id=int(produto)) & Q(location_id=int(local))
                & Q(date__gte=inicio) & Q(date__lte=final)))
        else:
            queryset = queryset[:10]
        return queryset
    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())

        # Calcular a média e o desvio padrão
        media = queryset.aggregate(media=Avg('price'))['media']
        desvio_padrao = queryset.aggregate(desvio_padrao=StdDev('price'))['desvio_padrao']
        minimo = queryset.aggregate(media=Min('price'))['media']
        maximo = queryset.aggregate(desvio_padrao=Max('price'))['desvio_padrao']

        # Serialize os dados
        serializer = self.get_serializer(queryset, many=True)
        
        # Inclua a média e o desvio padrão na resposta
        data = {}
        data['list'] = serializer.data
        data['media'] = media
        data['desvio_padrao'] = desvio_padrao
        data['minimo'] = minimo
        data['maximo'] = maximo

        return Response(data)

class LocationView(viewsets.ModelViewSet):
    queryset = Localizacao_Cotacao.objects.all()
    serializer_class = serLocations

class ProdutoView(viewsets.ModelViewSet):
    queryset = Commodity_Cotacao.objects.all()
    serializer_class = serProdutos

def parse_element_kml(element):
    coordinates = []
    for child in element.iterchildren():
        tag_name = etree.QName(child.tag).localname  # Get the local name of the tag
        # Handling Polygons
        if tag_name == 'Polygon':
            coords = str(child.outerBoundaryIs.LinearRing.coordinates).strip().split()
            for coord in coords:
                coordinates.append({'lat': float(coord.split(",")[1]), 'lng': float(coord.split(",")[0])})
            coordinates.pop() #retira último elemento da lista
        # Handling LineStrings
        elif tag_name == 'LineString':
            coords = str(child.coordinates).strip().split()
            for coord in coords:
                coordinates.append({'lat': float(coord.split(",")[1]), 'lng': float(coord.split(",")[0])})
        # Handling Points
        elif tag_name == 'Point':
            coords = str(child.coordinates).strip().split()
            for coord in coords:
                coordinates.append({'lat': float(coord.split(",")[1]), 'lng': float(coord.split(",")[0])})
        else:
            # Recursively handle complex elements like Placemark or Folder
            coordinates.extend(parse_element_kml(child))
    return coordinates

@csrf_exempt
def extract_coordinates_kml(request):
    #EXTRAIR COORDENADAS DE ARQUIVO KML
    context = {}
    def chunks(lst, n):
        """Split the list into chunks of a specified size."""
        for i in range(0, len(lst), n):
            yield lst[i:i + n]

    if request.method == "POST":
        file = request.FILES.get('file')
        if not file:
            return JsonResponse({'file':'Submeta um Arquivo'}, status=400)
        if not file.name.endswith('.kml'):
            return JsonResponse({'file':'O arquivo deve ser em formato KML'}, status=400)
        else:
            uploaded_file = request.FILES.get('file')
            uploaded_file.seek(0) 
            root = parser.parse(uploaded_file).getroot()
            coordinates_gd = parse_element_kml(root)
            list_lat_long = [(d['lat'], d['lng']) for d in coordinates_gd]
            all_elevations = []
            
            for chunk in chunks(list_lat_long, 400):
                elevations_chunk = Frasson.getElevationGoogleMapsAPI(locations=chunk)
                all_elevations.extend(elevations_chunk)

            coordinates = [{
                'number': index,
                'lat_gd': round(coord['location']['lat'], 6),
                'lat_gms': Frasson.convert_dd_to_dms(lat=coord['location']['lat']),
                'lat_utm': Frasson.convert_decimal_degrees_to_utm(lat=coord['location']['lat'], lon=coord['location']['lng'])['lat'],
                'lng_gd': round(coord['location']['lng'], 6),
                'lng_gms': Frasson.convert_dd_to_dms(long=coord['location']['lng']),
                'lng_utm': Frasson.convert_decimal_degrees_to_utm(lat=coord['location']['lat'], lon=coord['location']['lng'])['lng'],
                'elevation': int(coord['elevation'])
            }for index, coord in enumerate(all_elevations, start=1)]

            context['coordinates'] = coordinates
    return JsonResponse(context)

@csrf_exempt
def convert_html_table_to_excel(request):
    if request.method == "POST":
        #current timestamp
        time_now = int(time.time())
        #File name
        file_name = f"report_{time_now}"
        # Retrieve HTML content from POST data
        html_content = request.POST.get('html_content', '')
        # Use BeautifulSoup to parse the HTML
        soup = BeautifulSoup(html_content, 'html.parser')
        # Find the first table in the HTML
        table = soup.find('table')
        # Create a new Excel Workbook
        wb = Workbook()
        ws = wb.active  # Get active worksheet
        col_widths = {}  # A dictionary to hold the maximum length of each column
        # Center alignment style
        center_aligned_style = Alignment(horizontal='center', vertical='center')

        # Populate the worksheet with data from the HTML table
        for i, row in enumerate(table.find_all('tr')):
            for j, cell in enumerate(row.find_all(['td', 'th'])):
                cell_value = cell.get_text().strip()

                # Set the value and format for specific columns
                current_cell = ws.cell(row=i+1, column=j+1, value=cell_value)
                current_cell.alignment = center_aligned_style  # Apply center alignment

                try:
                    if j in [1, 2, 5, 6]:  # For columns 1, 2, 5 e 6
                        current_cell.value = float(cell_value)
                        current_cell.number_format = '0.00'
                    elif j in [7]:  # For column 5
                        current_cell.value = int(cell_value)
                        current_cell.number_format = '0'
                except ValueError:
                    pass  # Use text if conversion fails

                # Update column width (the length of the cell's content + 2 for some padding)
                col_widths[j] = max(col_widths.get(j, 0), len(cell_value) + 2)

        # Apply column widths
        for idx, width in col_widths.items():
            col_letter = ws.cell(row=1, column=idx+1).column_letter
            ws.column_dimensions[col_letter].width = width

        # Create a BytesIO object and save the workbook to it
        excel_file = BytesIO()
        wb.save(excel_file)
        # Set the pointer of BytesIO object to the beginning
        excel_file.seek(0)
        # Create an HttpResponse with the XLSX content
        response = HttpResponse(
            excel_file.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename={file_name}.xlsx'
        # Close the BytesIO object
        excel_file.close()
        return response
    else:
        return HttpResponse(status=404)

@csrf_exempt
def convert_html_table_to_excel_bnb(request):
    if request.method == "POST":
        #current timestamp
        time_now = int(time.time())
        #File name
        file_name = f"report_{time_now}"
        # Retrieve HTML content from POST data
        html_content = request.POST.get('html_content', '')
        # Use BeautifulSoup to parse the HTML
        soup = BeautifulSoup(html_content, 'html.parser')
        # Find the first table in the HTML
        table = soup.find('table')
        # Create a new Excel Workbook
        wb = Workbook()
        ws = wb.active  # Get active worksheet
        col_widths = {}  # A dictionary to hold the maximum length of each column
        # Center alignment style
        center_aligned_style = Alignment(horizontal='center', vertical='center')

        # Populate the worksheet with data from the HTML table
        for i, row in enumerate(table.find_all('tr')):
            k = 0
            for j, cell in enumerate(row.find_all(['td', 'th'])):
                cell_value = cell.get_text().strip()
                # Set the value and format for specific columns
                if j in [0,1,2,7]:
                    current_cell = ws.cell(row=i+1, column=k+1, value=cell_value)
                    current_cell.alignment = center_aligned_style  # Apply center alignment
                    k+=1
                try:
                    if j in [0]:
                        current_cell.value = float(cell_value)
                        current_cell.number_format = '0'
                    elif j in [1, 2, 7]:
                        current_cell.value = int(cell_value)
                        current_cell.number_format = '0.00'
                except ValueError:
                    pass  # Use text if conversion fails
                # Update column width (the length of the cell's content + 2 for some padding)
                col_widths[j] = max(col_widths.get(j, 0), len(cell_value) + 2)

        # Apply column widths
        for idx, width in col_widths.items():
            col_letter = ws.cell(row=1, column=idx+1).column_letter
            ws.column_dimensions[col_letter].width = width

        # Create a BytesIO object and save the workbook to it
        excel_file = BytesIO()
        wb.save(excel_file)
        # Set the pointer of BytesIO object to the beginning
        excel_file.seek(0)
        # Create an HttpResponse with the XLSX content
        response = HttpResponse(
            excel_file.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename={file_name}.xlsx'
        # Close the BytesIO object
        excel_file.close()
        return response
    else:
        return HttpResponse(status=404)
    
@csrf_exempt
def create_coordinates_pivots(request):
    #CRIAR COORDENADAS PIVOTS (LIMITE)
    context = {}
    if request.method == "POST":
        if (not (request.POST.get('latitude_gd') or request.POST.get('longitude_gd') or request.POST.get('quantidade') or request.POST.get('area_ha'))):
            error = {}
            if (not request.POST.get('latitude_gd')):
                error['latitude_gd'] = 'Campo Obrigatório'
            if (not request.POST.get('longitude_gd')):
                error['longitude_gd'] = 'Campo Obrigatório'
            if (not request.POST.get('quantidade')):
                error['quantidade'] = 'Campo Obrigatório'
            if (not request.POST.get('area_ha')):
                error['area_ha'] = 'Campo Obrigatório'    
            return JsonResponse(error, status=400)

        else:    
            latitude = float(request.POST.get('latitude_gd'))
            longitude = float(request.POST.get('longitude_gd'))
            area = float(request.POST.get('area_ha'))
            raio = math.sqrt(((area * 10000) / math.pi))
            quantidade_pontos = int(request.POST.get('quantidade'))
            coordinates_gd = Frasson.createLatLongPointsPivot(lat=latitude, lng=longitude, radius=raio, quant=quantidade_pontos)
            list_lat_long = [(d['lat'], d['lng']) for d in coordinates_gd]   
            resultado = Frasson.getElevationGoogleMapsAPI(locations=list_lat_long)

            coordinates = []
            for index, coord in enumerate(resultado, start=1):
                coordinates.append({
                    'number': index,
                    'lat_gd': round(coord['location']['lat'], 6),
                    'lat_gms': Frasson.convert_dd_to_dms(lat=coord['location']['lat']),
                    'lat_utm': Frasson.convert_decimal_degrees_to_utm(lat=coord['location']['lat'], lon=coord['location']['lng'])['lat'],
                    'lng_gd': round(coord['location']['lng'], 6),
                    'lng_gms': Frasson.convert_dd_to_dms(long=coord['location']['lng']),
                    'lng_utm': Frasson.convert_decimal_degrees_to_utm(lat=coord['location']['lat'], lon=coord['location']['lng'])['lng'],
                    'elevation': int(coord['elevation'])
                })

                context['coordinates'] = coordinates
    return JsonResponse(context)

@csrf_exempt
def download_kml_file_pivot(request):
    #DOWNLOAD OF KML FILE GENERATED FROM A CENTER PIVOT LIMIT AREA
    coordenadas = []
    if request.method == "POST":
        time_now = int(time.time()) #for the file name
        file_name = f"pivot_{time_now}.kml"
        # Retrieve HTML content from POST data
        html_content = request.POST.get('html_content', '')
        # Use BeautifulSoup to parse the HTML
        soup = BeautifulSoup(html_content, 'html.parser')
        # Find the first table in the HTML
        table = soup.find('table')

        # Populate the worksheet with data from the HTML table
        for i, row in enumerate(table.find_all('tr')):  # loop in table lines
            for j, cell in enumerate(row.find_all('td')):  # loop in columns of table lines
                if j == 1:  # column 2 (Latitude GD)
                    latitude = cell.get_text().strip().replace(',', '.')
                elif j == 2:  # column 3 (Longitude GD)
                    longitude = cell.get_text().strip().replace(',', '.')

            if i > 0:  # avoid the first line of the table (head)
                coordenadas.append({'lat': latitude, 'lng': longitude})

        # Create a list of coordinates for the polygon
        polygon_coordinates = []
        for coord in coordenadas:
            str_coordenada = f"{coord['lng']},{coord['lat']}"
            polygon_coordinates.append(str_coordenada)

        first_coordinate = f"{coordenadas[0]['lng']},{coordenadas[0]['lat']}" #first coordinate
        polygon_coordinates.append(first_coordinate)  #add again the first coordinate to close the polygon

        # Create a Polygon Style
        polygon_style = KML.Style(
            KML.id("polygon_style"),
            KML.PolyStyle(
                KML.color('00ffffff'),  # Fully transparent
                KML.fill(0),  # 0 means no fill
            ),
            KML.LineStyle(
                KML.color('ff00ff00'),  # Green outline
                KML.width(2),  # You can also set the line width
            )
        )

        # Create a KML Document with a Polygon and Style
        kml_doc = KML.kml(
            KML.Document(
                polygon_style,  # Add the Style to the Document
                KML.Folder(
                    KML.name("Polygon"),
                    KML.Placemark(
                        KML.name("Pivot"),
                        KML.styleUrl("#polygon_style"),  # Refer to the polygon_style by id
                        KML.Polygon(
                            KML.outerBoundaryIs(
                                KML.LinearRing(
                                    KML.coordinates(" ".join(polygon_coordinates))
                                )
                            )
                        )
                    )
                )
            )
        )

        # Convert KML document to string
        kml_str = etree.tostring(kml_doc, pretty_print=True)
        # Create a response with the KML type
        response = HttpResponse(kml_str, content_type='application/vnd.google-earth.kml+xml')
        # Add a file attachment header
        response['Content-Disposition'] = f'attachment; filename="{file_name}"'
        return response
    else:
        return JsonResponse(status=404)

@csrf_exempt
def download_kml_polygon_coordinates(request):
    if request.method == "POST":
        time_now = int(time.time()) #for the file name
        file_name = f"pivot_{time_now}.kml"
        data = json.loads(request.body.decode('utf-8'))

        polygon_coordinates = []
        for coord in data[0]["points"]:
            str_coordenada = f"{coord['lng']},{coord['lat']}"
            polygon_coordinates.append(str_coordenada)

        first_coordinate = f"{data[0]['points'][0]['lng']},{data[0]['points'][0]['lat']}" #first coordinate
        polygon_coordinates.append(first_coordinate)  #add again the first coordinate to close the polygon

        # Create a Polygon Style
        polygon_style = KML.Style(
            KML.id("polygon_style"),
            KML.PolyStyle(
                KML.color('00ffffff'),  # Fully transparent
                KML.fill(0),  # 0 means no fill
            ),
            KML.LineStyle(
                KML.color('ff00ff00'),  # Green outline
                KML.width(2),  # You can also set the line width
            )
        )
        # Create a KML Document with a Polygon and Style
        kml_doc = KML.kml(
            KML.Document(
                polygon_style,  # Add the Style to the Document
                KML.Folder(
                    KML.name("Polygon"),
                    KML.Placemark(
                        KML.name("Polygon"),
                        KML.styleUrl("#polygon_style"),  # Refer to the polygon_style by id
                        KML.Polygon(
                            KML.outerBoundaryIs(
                                KML.LinearRing(
                                    KML.coordinates(" ".join(polygon_coordinates))
                                )
                            )
                        )
                    )
                )
            )
        )
        # Convert KML document to string
        kml_str = etree.tostring(kml_doc, pretty_print=True)
        # Create a response with the KML type
        response = HttpResponse(kml_str, content_type='application/vnd.google-earth.kml+xml')
        # Add a file attachment header
        response['Content-Disposition'] = f'attachment; filename="{file_name}"'
        return response


@csrf_exempt
def download_kml_point_coordinates(request):
    data = json.loads(request.body.decode('utf-8'))
    # Create a KML document with a Folder
    doc = KML.kml(
        KML.Document(
            KML.Folder(
                KML.name("Coordenadas")
            )
        )
    )
    # Retrieve the Folder element
    folder = doc.Document.Folder
    coordinates = []
    for coord in data["coordinates"]:
        str_coord = f"{coord['lng']},{coord['lat']}" 
        coordinates.append((str_coord))

    # Add a Placemark for each coordinate
    for i, coordinate in enumerate(coordinates):
        placemark = KML.Placemark(
            KML.name(i + 1),
            KML.Point(
                KML.coordinates(coordinate)
            )
        )
        folder.append(placemark)
    # Convert to string
    kml_str = etree.tostring(doc, pretty_print=True)
    # Create a response with the KML type
    response = HttpResponse(kml_str, content_type='application/vnd.google-earth.kml+xml')
    # Add a file attachment header
    response['Content-Disposition'] = f'attachment; filename="kml_file"'
    return response


@csrf_exempt
def kml_transform_polygon(request):
    context = {}
    if request.method == "POST":
        file = request.FILES.get('file')
        if not file:
            return JsonResponse({'file':'Submeta um Arquivo'}, status=400)
        if not file.name.endswith('.kml'):
            return JsonResponse({'file':'O arquivo deve ser em formato KML'}, status=400)
        else:
            uploaded_file = request.FILES['file']
            uploaded_file.seek(0)
            root = parser.parse(uploaded_file).getroot()
            coordinates_gd = parse_element_kml(root)  # Sua função customizada para extrair coordenadas
            coordinates_gd.append(coordinates_gd[0]) #add o primeiro registro ao final para fechar o poligono

            coordinates = [{
                'lat': round(coord['lat'], 6),
                'lng': round(coord['lng'], 6),
            }for coord in coordinates_gd]
            context['coordinates'] = coordinates
    return JsonResponse(context)


@csrf_exempt
def download_kml_transform_polygon(request):
    #DOWNLOAD DO ARQUIVO KML NO NOVO FORMATO PARA PLATAFORMA BB
    if request.method == "POST":
        time_now = int(time.time()) #for the file name
        file_name = f"pivot_{time_now}.kml"
        data = json.loads(request.body.decode('utf-8'))
        polygon_coordinates = [f"{coord['lng']},{coord['lat']}" for coord in data["coordinates"]]
        first_coordinate = f"{data['coordinates'][0]['lng']},{data['coordinates'][0]['lat']}" #first coordinate
        polygon_coordinates.append(first_coordinate)  #add again the first coordinate to close the polygon
        # Define o namespace principal e os prefixos
        NSMAP = {
            None: "http://www.opengis.net/kml/2.2",
            'gx': "http://www.google.com/kml/ext/2.2",
            'atom': "http://www.w3.org/2005/Atom"
        }
        # Cria o elemento raiz KML
        kml = etree.Element("kml", nsmap=NSMAP)
        # Adiciona o Documento ao KML
        document = etree.SubElement(kml, "Document")
        # Adiciona o nome ao Documento
        name = etree.SubElement(document, "name")
        name.text = file_name
        # Cria um StyleMap
        style_map = etree.SubElement(document, "StyleMap", id="msn_ylw-pushpin")
        pairs = [
            ("normal", "#sn_ylw-pushpin"),
            ("highlight", "#sh_ylw-pushpin")
        ]
        for key, url in pairs:
            pair = etree.SubElement(style_map, "Pair")
            key_elem = etree.SubElement(pair, "key")
            key_elem.text = key
            style_url = etree.SubElement(pair, "styleUrl")
            style_url.text = url

        styles = [
            ("sn_ylw-pushpin", 1.1),
            ("sh_ylw-pushpin", 1.3)
        ]

        for id, scale in styles:
            style = etree.SubElement(document, "Style", id=id)
            icon_style = etree.SubElement(style, "IconStyle")
            scale_elem = etree.SubElement(icon_style, "scale")
            scale_elem.text = str(scale)
            icon = etree.SubElement(icon_style, "Icon")
            href = etree.SubElement(icon, "href")
            href.text = "http://maps.google.com/mapfiles/kml/pushpin/ylw-pushpin.png"
            hot_spot = etree.SubElement(icon_style, "hotSpot", x="20", y="2", xunits="pixels", yunits="pixels")
        # Adiciona um Placemark
        placemark = etree.SubElement(document, "Placemark")
        placemark_name = etree.SubElement(placemark, "name")
        placemark_name.text = file_name
        style_url = etree.SubElement(placemark, "styleUrl")
        style_url.text = "#msn_ylw-pushpin"
        polygon = etree.SubElement(placemark, "Polygon")
        tessellate = etree.SubElement(polygon, "tessellate")
        tessellate.text = "1"
        outer_boundary_is = etree.SubElement(polygon, "outerBoundaryIs")
        linear_ring = etree.SubElement(outer_boundary_is, "LinearRing")
        coordinates = etree.SubElement(linear_ring, "coordinates")
        coordinates.text = " ".join(polygon_coordinates)
        # Converte para string
        kml_str = etree.tostring(kml, pretty_print=True, xml_declaration=True, encoding='UTF-8')
        # Create a response with the KML type
        response = HttpResponse(kml_str, content_type='application/vnd.google-earth.kml+xml')
        response['Content-Disposition'] = f'attachment; filename="{file_name}"'
        return response