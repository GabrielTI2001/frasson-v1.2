from django.shortcuts import render, HttpResponse
from django.db.models import Q
from django.http import JsonResponse
from rest_framework import viewsets
from .serializers import ListRegimes, detailRegimes, listFarms, detailFarms
from pipefy.models import Regimes_Exploracao, Imoveis_Rurais, Operacoes_Contratadas
from io import BytesIO
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, NamedStyle
import time, json
from django.views.decorators.csrf import csrf_exempt

class RegimesView(viewsets.ModelViewSet):
    queryset = Regimes_Exploracao.objects.all()
    serializer_class = detailRegimes
    def get_queryset(self):
        queryset = super().get_queryset()
        cliente_term = self.request.query_params.get('cliente', None)   
        instituicao_term = self.request.query_params.get('instituicao', None)  
        if cliente_term and instituicao_term:
            queryset = queryset.filter(quem_explora_id=int(cliente_term), instituicao_id=int(instituicao_term))
        return queryset
    def get_serializer_class(self):
        if self.action == 'list':
            return ListRegimes
        else:
            return self.serializer_class

class FarmsView(viewsets.ModelViewSet):
    queryset = Imoveis_Rurais.objects.all()
    serializer_class = detailFarms
    # permission_classes = [IsAuthenticated]
    lookup_field = 'id'
    def get_queryset(self):
        queryset = super().get_queryset()
        search_term = self.request.query_params.get('search', None)
        all_term = self.request.query_params.get('all', None)
        if search_term:
            queryset = queryset.filter(
                Q(nome_imovel__icontains=search_term)
            )
        elif all_term:
            queryset = queryset.order_by('-created_at')
        else:
            if self.action == 'list':
                queryset = queryset.order_by('-created_at')[:10]
        return queryset
    def get_serializer_class(self):
        if self.action == 'list':
            return listFarms
        else:
            return self.serializer_class
        
def creditData(request):
    anos_operacoes = [y['data_emissao_cedula__year'] for y in Operacoes_Contratadas.objects.values('data_emissao_cedula__year').distinct()]
    instituicoes = Operacoes_Contratadas.objects.values('instituicao__instituicao__id','instituicao__instituicao__razao_social').distinct()
    lista_instituicoes = [{
        'id': instituicao['instituicao__instituicao__id'],
        'instituicao': instituicao['instituicao__instituicao__razao_social']
    }for instituicao in instituicoes]

    data = {
        'anos': anos_operacoes,
        'instituicoes': lista_instituicoes,
    }
    return JsonResponse(data)

@csrf_exempt
def convert_html_table_to_excel(request):
    if request.method == "POST":
        time_now = int(time.time())
        file_name = f"report_{time_now}"
        request_data = json.loads(request.body)
        html_content = request_data.get('html_content', '')
        soup = BeautifulSoup(html_content, 'html.parser')
        table = soup.find('table')

        # Cria um Excel Workbook
        wb = Workbook()
        ws = wb.active  # Get active worksheet

        col_widths = {}  # A dictionary to hold the maximum length of each column
        center_aligned_style = Alignment(horizontal='center', vertical='center')
        for i, row in enumerate(table.find_all('tr')):
            for j, cell in enumerate(row.find_all(['td', 'th'])):
                cell_value = cell.get_text().strip()
                # Prrenche colunas do excel com valores da table
                current_cell = ws.cell(row=i+1, column=j+1, value=cell_value)
                current_cell.alignment = center_aligned_style 
                col_widths[j] = max(col_widths.get(j, 0), len(cell_value) + 2)

        # Apply column widths
        for idx, width in col_widths.items():
            col_letter = ws.cell(row=1, column=idx+1).column_letter
            ws.column_dimensions[col_letter].width = width

        # Create a BytesIO object and save the workbook to it
        excel_file = BytesIO()
        wb.save(excel_file)
        # Set the pointer of BytesIO object to the beginning
        excel_file.seek(0)

        response = HttpResponse(
            excel_file.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename={file_name}.xlsx'
        excel_file.close()

        return response
    else:
        return HttpResponse(404)